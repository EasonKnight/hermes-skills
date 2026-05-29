#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
校对脚本：对比程序生成的合并月报 vs 交易所官方月结算单PDF
用法: python verify_monthly_report.py
"""

import re, sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    import pdfplumber
except ImportError as e:
    print(f"缺少依赖: {e}")
    sys.exit(1)

# ── 配置 ──────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
XLSX_PATH = SCRIPT_DIR / "R202605_合并月报.xlsx"
PDF_PATH = SCRIPT_DIR / "900002266恒力石化销售（海口）有限公司_20260501_20260529(1).pdf"


def to_float(s):
    if s is None: return 0.0
    s = str(s).replace(',', '').replace('\xa0', '').replace('--', '0').strip()
    if not s or s == '-' or s == '---': return 0.0
    try: return float(s)
    except ValueError: return 0.0


def clean(s):
    if s is None: return ''
    return str(s).replace('\xa0', '').replace('\n', '').replace('\r', '').strip()


# ═══════════════════════════════════════════════════════════════
# 1. 解析程序生成的 xlsx
# ═══════════════════════════════════════════════════════════════

def parse_generated_xlsx(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    result = {
        'funds': {},       # 资金状况
        'deposits': [],    # 出入金
        'fees': [],        # 手续费
        'futures': [],     # 期货成交
        'options': [],     # 期权成交
    }

    # 解析资金状况 — 动态搜索A/F列标题
    name_to_key = {
        '上月结存': '上月结存', '当月存取合计': '当月存取合计',
        '当月盈亏': '当月盈亏', '当月总权利金': '当月总权利金',
        '当月手续费': '当月手续费', '当月结存': '当月结存',
        '浮动盈亏': '浮动盈亏', '客户权益': '客户权益',
        '实有货币资金': '实有货币资金', '保证金占用': '保证金占用',
        '可用资金': '可用资金', '风险度': '风险度',
        '追加保证金': '追加保证金',
    }
    for row in ws.iter_rows(min_row=7, max_row=15, values_only=False):
        a_val = clean(row[0].value)
        c_val = row[2].value
        f_val = clean(row[5].value)
        h_val = row[7].value

        if a_val in name_to_key and c_val is not None:
            result['funds'][name_to_key[a_val]] = to_float(c_val) if '%' not in str(c_val) else clean(c_val)
        if f_val in name_to_key and h_val is not None:
            result['funds'][name_to_key[f_val]] = to_float(h_val) if '%' not in str(h_val) else clean(h_val)

    # 解析出入金
    in_deposits = False
    for row in ws.iter_rows(min_row=17, max_row=ws.max_row, values_only=False):
        a = clean(row[0].value)
        if '出入金明细' in a:
            in_deposits = True; continue
        if in_deposits and ('其它资金' in a or '期货成交' in a):
            break
        if not in_deposits: continue
        if a == '合计' or a == '发生日期': continue
        if not a: continue
        dep_in = to_float(row[1].value)
        dep_out = to_float(row[2].value)
        if dep_in > 0 or dep_out > 0:
            result['deposits'].append({'date': a, 'in': dep_in, 'out': dep_out})

    # 解析手续费
    in_fees = False
    for row in ws.iter_rows(min_row=20, max_row=ws.max_row, values_only=False):
        a = clean(row[0].value)
        if '其它资金明细' in a:
            in_fees = True; continue
        if in_fees and ('期货成交' in a):
            break
        if not in_fees: continue
        if a == '合计' or not a: continue
        d_val = to_float(row[3].value)
        if d_val != 0:
            result['fees'].append({
                'date': a, 'exchange': clean(row[1].value),
                'type': clean(row[2].value), 'amount': d_val,
            })

    # 解析期货成交 — 动态查找区域
    futures_start = None
    options_start = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a = clean(row[0].value)
        if '期货成交汇总' in a:
            futures_start = row[0].row
        if '期权成交汇总' in a:
            options_start = row[0].row
            break

    in_futures = False
    if futures_start:
        for row in ws.iter_rows(min_row=futures_start, max_row=ws.max_row, values_only=False):
            a = clean(row[0].value)
            if '期货成交汇总' in a:
                in_futures = True; continue
            if in_futures and ('期权成交' in a):
                break
            if not in_futures: continue
            if a == '合计' or a == '交易日期': continue
            if not a: continue
            result['futures'].append({
                'date': a,
                'contract': clean(row[1].value),
                'bs': clean(row[2].value),
                'lots': int(to_float(row[5].value)),
                'fee': to_float(row[8].value),
                'pnl': to_float(row[9].value) if '--' not in str(row[9].value or '') else 0.0,
            })

    # 解析期权成交
    in_options = False
    if options_start:
        for row in ws.iter_rows(min_row=options_start, max_row=ws.max_row, values_only=False):
            a = clean(row[0].value)
            if '期权成交汇总' in a:
                in_options = True; continue
            if not in_options: continue
            if a == '合计' or a == '日期': continue
            if '按合同约定' in a: break
            if not a: continue
            result['options'].append({
                'date': a,
                'contract': clean(row[1].value),
                'bs': clean(row[5].value),
                'lots': int(to_float(row[7].value)),
                'premium': to_float(row[8].value),
                'fee': to_float(row[10].value),
            })

    wb.close()
    return result


# ═══════════════════════════════════════════════════════════════
# 2. 解析交易所官方 PDF
# ═══════════════════════════════════════════════════════════════

def parse_official_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        full_text = '\n'.join(p.extract_text() or '' for p in pdf.pages)

    result = {
        'funds': {},
        'trades': [],
    }

    # ── 资金状况 ──
    field_map = {
        '期初结存': '上月结存', '出 入 金': '当月存取合计',
        '平仓盈亏': '当月盈亏', '浮动盈亏': '浮动盈亏',
        '手 续 费': '当月手续费', '期末结存': '当月结存',
        '客户权益': '客户权益', '保证金占用': '保证金占用',
        '可用资金': '可用资金', '风险度': '风险度',
        '应追加资金': '追加保证金', '权利金收入': '当月总权利金',
        '市值': '市值',
    }
    for pdf_field, our_key in field_map.items():
        m = re.search(rf'{re.escape(pdf_field)}:\s*([\-\d,]+\.?\d*%?)', full_text)
        if m:
            val = m.group(1)
            if '%' in val:
                result['funds'][our_key] = val
            else:
                result['funds'][our_key] = to_float(val)

    # ── 成交记录 ──
    in_trades = False
    trade_lines = []
    for line in full_text.split('\n'):
        if '成交记录' in line:
            in_trades = True; continue
        if in_trades and ('持仓明细' in line or '持仓汇总' in line):
            break
        if not in_trades: continue
        line = line.strip()
        if not line or line.startswith('-') or line.startswith('成交日期'): continue
        if re.match(r'第\s*\d+\s*页', line): continue
        trade_lines.append(line)

    # 拼接被分行的记录
    merged = []
    current = ''
    for line in trade_lines:
        if re.match(r'^\d{8}\s', line):
            if current:
                merged.append(current)
            current = line
        else:
            current += ' ' + line
    if current:
        merged.append(current)

    for line in merged:
        parts = line.split()
        if len(parts) < 11: continue
        if not re.match(r'^\d{8}$', parts[0]): continue

        date_8d = parts[0]
        contract = parts[3]
        bs = parts[4]

        # 跳过聚合平仓汇总行 (bs='-')
        if bs == '-':
            continue

        # 检测是否为期权
        is_option = ('期权' in parts[2] or
                     bool(re.search(r'[CP]\d{3,}', contract.upper())) or
                     bool(re.search(r'-[CP]-', contract.upper())))

        lots = int(to_float(parts[7]))
        oc = parts[9] if len(parts) > 9 else ''
        fee = to_float(parts[10]) if len(parts) > 10 else 0
        pnl = 0.0
        premium = 0.0

        if is_option:
            premium = to_float(parts[11]) if len(parts) > 11 else 0
        else:
            pnl = to_float(parts[11]) if len(parts) > 11 else 0

        result['trades'].append({
            'date': f'{date_8d[:4]}-{date_8d[4:6]}-{date_8d[6:]}',
            'contract': contract,
            'bs': bs,
            'lots': lots,
            'oc': oc,
            'fee': fee,
            'pnl': pnl,
            'premium': premium,
            'is_option': is_option,
        })

    return result


# ═══════════════════════════════════════════════════════════════
# 3. 对比分析
# ═══════════════════════════════════════════════════════════════

def compare_funds(gen, off):
    """对比资金状况"""
    print("=" * 70)
    print("  一、资金状况对比")
    print("=" * 70)
    print(f"  {'项目':<16} {'程序生成':>14} {'交易所官方':>14} {'差异':>14}")

    keys = ['上月结存', '当月存取合计', '当月盈亏', '当月总权利金',
            '当月手续费', '当月结存', '浮动盈亏', '客户权益',
            '保证金占用', '可用资金', '风险度', '追加保证金']
    errors = []
    for key in keys:
        gv = gen.get(key, 0)
        ov = off.get(key, 0)
        if isinstance(gv, str) or isinstance(ov, str):
            match = str(gv) == str(ov)
            marker = '✅' if match else '❌'
            diff = str(gv) if not match else ''
        else:
            diff = gv - ov
            marker = '✅' if abs(diff) < 0.02 else '❌'
        print(f"  {key:<16} {str(gv):>14} {str(ov):>14} {marker} {diff if diff else ''}")
        if marker == '❌':
            errors.append(key)
    return errors


def compare_trades_daily(gen, off):
    """按日期对比手续费和盈亏"""
    print()
    print("=" * 70)
    print("  二、每日手续费 & 平仓盈亏对比（按日汇总）")
    print("=" * 70)

    gen_by_date = defaultdict(lambda: {'fee': 0.0, 'pnl': 0.0, 'premium': 0.0})
    off_by_date = defaultdict(lambda: {'fee': 0.0, 'pnl': 0.0, 'premium': 0.0})

    for t in gen['futures']:
        d = t['date']
        gen_by_date[d]['fee'] += t['fee']
        gen_by_date[d]['pnl'] += t['pnl']
    for t in gen['options']:
        d = t['date']
        gen_by_date[d]['fee'] += t['fee']
        gen_by_date[d]['premium'] += abs(t.get('premium', 0))

    for t in off['trades']:
        d = t['date']
        off_by_date[d]['fee'] += t['fee']
        if t['is_option']:
            off_by_date[d]['premium'] += abs(t['premium'])
        else:
            off_by_date[d]['pnl'] += t['pnl']

    all_dates = sorted(set(list(gen_by_date.keys()) + list(off_by_date.keys())))

    print(f"  {'日期':<12} {'生成费':>10} {'官方费':>10} {'费差':>10} | {'生成盈亏':>10} {'官方盈亏':>10} {'盈亏差':>10}")
    print(f"  {'-'*78}")
    gen_total_fee = off_total_fee = 0.0
    gen_total_pnl = off_total_pnl = 0.0
    date_errors = []

    for d in all_dates:
        gf = gen_by_date[d]['fee']
        of_ = off_by_date[d]['fee']
        gp = gen_by_date[d]['pnl'] + gen_by_date[d]['premium']
        op = off_by_date[d]['pnl'] + off_by_date[d]['premium']
        fd = gf - of_
        pd = gp - op
        gen_total_fee += gf
        off_total_fee += of_
        gen_total_pnl += gp
        off_total_pnl += op
        fee_ok = abs(fd) < 0.02
        pnl_ok = abs(pd) < 0.02
        marker = '✅' if fee_ok and pnl_ok else '❌'
        print(f"  {d:<12} {gf:>10.2f} {of_:>10.2f} {fd:>+10.2f} | {gp:>10.0f} {op:>10.0f} {pd:>+10.0f} {marker}")
        if not fee_ok or not pnl_ok:
            date_errors.append(d)

    print(f"  {'-'*78}")
    tf = gen_total_fee - off_total_fee
    tp = gen_total_pnl - off_total_pnl
    print(f"  {'合计':<12} {gen_total_fee:>10.2f} {off_total_fee:>10.2f} {tf:>+10.2f} | {gen_total_pnl:>10.0f} {off_total_pnl:>10.0f} {tp:>+10.0f}")

    return date_errors


def compare_trade_detail(gen, off):
    """按 (日期, 合约, 方向) 聚合对比"""
    print()
    print("=" * 70)
    print("  三、成交明细聚合对比（按 日期+合约+方向）")
    print("=" * 70)

    gen_agg = defaultdict(lambda: {'lots': 0, 'fee': 0.0, 'pnl': 0.0, 'premium': 0.0})
    for t in gen['futures']:
        key = (t['date'], t['contract'].upper(), t['bs'])
        gen_agg[key]['lots'] += t['lots']
        gen_agg[key]['fee'] += t['fee']
        gen_agg[key]['pnl'] += t.get('pnl', 0)
    for t in gen['options']:
        key = (t['date'], t['contract'].upper(), t['bs'])
        gen_agg[key]['lots'] += t['lots']
        gen_agg[key]['fee'] += t['fee']
        gen_agg[key]['premium'] += abs(t.get('premium', 0))

    off_agg = {}
    for t in off['trades']:
        key = (t['date'], t['contract'].upper(), t['bs'])
        if key not in off_agg:
            off_agg[key] = {'lots': 0, 'fee': 0.0, 'pnl': 0.0, 'premium': 0.0, 'is_option': t['is_option']}
        off_agg[key]['lots'] += t['lots']
        off_agg[key]['fee'] += t['fee']
        if t['is_option']:
            off_agg[key]['premium'] += abs(t['premium'])
        else:
            off_agg[key]['pnl'] += t['pnl']

    gen_keys = set(gen_agg.keys())
    off_keys = set(off_agg.keys())

    only_gen = gen_keys - off_keys
    only_off = off_keys - gen_keys
    both = gen_keys & off_keys

    if only_gen:
        print(f"\n  ⚠️ 仅程序有（{len(only_gen)}组）：")
        for k in sorted(only_gen):
            t = gen_agg[k]
            print(f"    {k[0]} {k[1]} {k[2]} {t['lots']}手 费{t['fee']:.2f}")

    if only_off:
        print(f"\n  ⚠️ 仅官方有（{len(only_off)}组）：")
        for k in sorted(only_off):
            t = off_agg[k]
            opt_tag = ' [期权]' if t['is_option'] else ''
            print(f"    {k[0]} {k[1]} {k[2]} {t['lots']}手 费{t['fee']:.2f}{opt_tag}")

    lot_mismatch = []
    fee_mismatch = []
    for k in sorted(both):
        gt = gen_agg[k]
        ot = off_agg[k]
        if gt['lots'] != ot['lots']:
            lot_mismatch.append((k, gt['lots'], ot['lots']))
        if abs(gt['fee'] - ot['fee']) > 0.05:
            fee_mismatch.append((k, gt['fee'], ot['fee']))

    if lot_mismatch:
        print(f"\n  ⚠️ 手数不一致（{len(lot_mismatch)}组）：")
        for k, gl, ol in lot_mismatch[:15]:
            print(f"    {k[0]} {k[1]} {k[2]} 生成{gl}手 官方{ol}手")

    if fee_mismatch:
        print(f"\n  ⚠️ 手续费不一致（{len(fee_mismatch)}组）：")
        for k, gf, of in fee_mismatch[:15]:
            print(f"    {k[0]} {k[1]} {k[2]} 生成费{gf:.2f} 官方费{of:.2f} 差{gf-of:+.2f}")

    matched = len(both) - len(lot_mismatch) - len(fee_mismatch)
    print(f"\n  ✅ 完全一致: {matched} 组 / 共{len(both)}组")
    print(f"  生成{len(gen_agg)}组, 官方{len(off_agg)}组, 共同{len(both)}组")

    return len(only_gen) + len(only_off)


def check_balance_equation(off):
    """验证官方数据自洽"""
    print()
    print("=" * 70)
    print("  四、官方数据公式验证")
    print("=" * 70)
    begin = off['funds'].get('上月结存', 0)
    deposit = off['funds'].get('当月存取合计', 0)
    pnl = off['funds'].get('当月盈亏', 0)
    fee = off['funds'].get('当月手续费', 0)
    premium = off['funds'].get('当月总权利金', 0)
    end = off['funds'].get('当月结存', 0)

    calculated = begin + deposit + pnl - fee + premium
    diff = end - calculated

    print(f"  期初结存:   {begin:>14.2f}")
    print(f"  + 出入金:   {deposit:>14.2f}")
    print(f"  + 平仓盈亏: {pnl:>14.2f}")
    print(f"  - 手续费:   {fee:>14.2f}")
    print(f"  + 权利金:   {premium:>14.2f}")
    print(f"  {'─'*30}")
    print(f"  = 计算结存: {calculated:>14.2f}")
    print(f"  实际结存:   {end:>14.2f}")
    print(f"  差异:       {diff:>14.2f} {'✅' if abs(diff) < 0.02 else '❌'}")


def main():
    print("🔍 月报校对工具")
    print(f"  程序生成: {XLSX_PATH.name}")
    print(f"  交易所官方: {PDF_PATH.name}")
    print()

    if not XLSX_PATH.exists():
        print(f"❌ 找不到: {XLSX_PATH}")
        return
    if not PDF_PATH.exists():
        print(f"❌ 找不到: {PDF_PATH}")
        return

    print("正在解析程序生成的月报...")
    gen = parse_generated_xlsx(XLSX_PATH)
    print(f"  期货{len(gen['futures'])}笔, 期权{len(gen['options'])}笔, "
          f"出入金{len(gen['deposits'])}笔, 手续费{len(gen['fees'])}条")

    print("正在解析交易所官方月结算单...")
    off = parse_official_pdf(PDF_PATH)
    futures_cnt = sum(1 for t in off['trades'] if not t['is_option'])
    options_cnt = sum(1 for t in off['trades'] if t['is_option'])
    print(f"  总成交{len(off['trades'])}笔 (期货{futures_cnt}, 期权{options_cnt}), "
          f"出入金{len(off.get('deposits', []))}笔")
    print(f"  资金字段: {len(off['funds'])}个")

    print()
    fund_errors = compare_funds(gen['funds'], off['funds'])
    date_errors = compare_trades_daily(gen, off)
    compare_trade_detail(gen, off)
    check_balance_equation(off)

    # ── 总结 ──
    print()
    print("=" * 70)
    print("  总结")
    print("=" * 70)

    if not fund_errors:
        print(f"\n  ✅ 资金状况全部12项一致")
    else:
        print(f"\n  ❌ 资金状况有 {len(fund_errors)} 项不一致: {', '.join(fund_errors)}")

    if not date_errors:
        print(f"  ✅ 每日手续费+盈亏汇总全部一致")
    else:
        print(f"  ❌ {len(date_errors)} 天的手续费/盈亏有差异")

    if fund_errors or date_errors:
        print(f"\n  ⚠️ 需人工核查以上差异项")
    else:
        print(f"\n  🎉 程序生成的月报与交易所官方月结算单数据一致！")


if __name__ == '__main__':
    main()
